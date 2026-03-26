#!/usr/bin/env python3
"""Export tasks/<folder>/test_cases.json to CSV and XLSX in the same folder."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path


def steps_to_text(steps: list) -> str:
    lines = []
    for i, s in enumerate(steps, start=1):
        lines.append(f"{i}. {s}")
    return "\n".join(lines)


def export_csv(path: Path, data: dict) -> None:
    cases = data.get("test_cases") or []
    out = path.with_suffix(".csv")
    fieldnames = [
        "jira_key",
        "pack_title",
        "prerequisites",
        "id",
        "category",
        "title",
        "test_type",
        "preconditions",
        "steps",
        "expected_result",
    ]
    prereq_joined = "\n\n".join(data.get("prerequisites") or [])
    jira = data.get("jira_key", "")
    pack_title = data.get("title", "")

    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for i, c in enumerate(cases):
            w.writerow(
                {
                    "jira_key": jira,
                    "pack_title": pack_title,
                    "prerequisites": prereq_joined if i == 0 else "",
                    "id": c.get("id", ""),
                    "category": c.get("category", ""),
                    "title": c.get("title", ""),
                    "test_type": c.get("test_type", ""),
                    "preconditions": c.get("preconditions", ""),
                    "steps": steps_to_text(c.get("steps") or []),
                    "expected_result": c.get("expected_result", ""),
                }
            )
    print(f"Wrote {out}")


def export_xlsx(path: Path, data: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    cases = data.get("test_cases") or []
    # Prefer README-style name: tasks/<folder>/<folder>.xlsx
    out = path.parent / f"{path.parent.name}.xlsx"

    wb = Workbook()
    info = wb.active
    info.title = "Pack info"
    info["A1"] = "jira_key"
    info["B1"] = data.get("jira_key", "")
    info["A2"] = "title"
    info["B2"] = data.get("title", "")
    info["A4"] = "prerequisites"
    info["B4"] = "\n\n".join(data.get("prerequisites") or [])
    info["A1"].font = info["A2"].font = info["A4"].font = Font(bold=True)
    for cell in ("B2", "B4"):
        info[cell].alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90

    tc = wb.create_sheet("Test cases", 1)
    headers = [
        "id",
        "category",
        "title",
        "test_type",
        "preconditions",
        "steps",
        "expected_result",
    ]
    for col, h in enumerate(headers, start=1):
        cell = tc.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    wrap = Alignment(wrap_text=True, vertical="top")
    for r, c in enumerate(cases, start=2):
        tc.cell(row=r, column=1, value=c.get("id", ""))
        tc.cell(row=r, column=2, value=c.get("category", ""))
        tc.cell(row=r, column=3, value=c.get("title", ""))
        tc.cell(row=r, column=4, value=c.get("test_type", ""))
        tc.cell(row=r, column=5, value=c.get("preconditions", ""))
        tc.cell(row=r, column=6, value=steps_to_text(c.get("steps") or []))
        tc.cell(row=r, column=7, value=c.get("expected_result", ""))
        for col in range(1, 8):
            tc.cell(row=r, column=col).alignment = wrap

    tc.freeze_panes = "A2"
    tc.column_dimensions["A"].width = 22
    tc.column_dimensions["B"].width = 22
    tc.column_dimensions["C"].width = 42
    tc.column_dimensions["D"].width = 12
    tc.column_dimensions["E"].width = 48
    tc.column_dimensions["F"].width = 52
    tc.column_dimensions["G"].width = 52

    wb.save(out)
    print(f"Wrote {out}")


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument(
        "json_path",
        type=Path,
        nargs="?",
        default=None,
        help="Path to test_cases.json (default: tasks/KUDO_35_family-members-list/test_cases.json under repo root)",
    )
    args = p.parse_args()
    root = Path(__file__).resolve().parent.parent
    json_path = args.json_path or (
        root / "tasks" / "KUDO_35_family-members-list" / "test_cases.json"
    )
    if not json_path.is_file():
        print(f"Missing {json_path}", file=sys.stderr)
        return 1
    data = json.loads(json_path.read_text(encoding="utf-8"))
    export_csv(json_path, data)
    try:
        export_xlsx(json_path, data)
    except ImportError:
        print("openpyxl not installed; pip install -r scripts/requirements.txt", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
